# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
from openpyxl import load_workbook
from io import BytesIO


class TediAllowanceImportWizard(models.TransientModel):
    _name = 'tedi.allowance.import.wizard'
    _description = 'Import Phụ cấp từ Excel'

    file = fields.Binary(string="Tệp Excel", required=True)
    filename = fields.Char(string="Tên file")

    def action_import(self):
        if not self.file:
            raise UserError(_("Vui lòng chọn file Excel!"))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Không thể đọc file. Vui lòng kiểm tra định dạng .xlsx. Chi tiết: %s") % str(e))

        Allowance = self.env['hr.employee.allowance']
        AllowanceType = self.env['hr.allowance.type']  # Model danh mục mới
        Employee = self.env['hr.employee']

        success_count = 0
        errors = []
        row_index = 1

        # Cache danh sách loại phụ cấp để tối ưu hiệu năng (Tránh query DB trong vòng lặp)
        # Dictionary dạng: {'MÃ_EXCEL': ID_DATABASE}
        # Lưu ý: Mã trong Excel so với mã trong DB nên chuẩn hóa (vd: upper(), strip())
        allowance_types_map = {
            rec.code: rec.id for rec in AllowanceType.search([])
        }

        # In ra để debug nếu cần
        # print("Available Types:", allowance_types_map)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1
            if not row or len(row) < 5 or row[4] is None:
                continue

            try:
                # 1. Lấy dữ liệu thô
                raw_month = row[0]
                raw_year = row[1]
                raw_type_code = str(row[2]).strip() if row[2] else False  # Đây là Mã loại (VD: AN_CA)
                raw_code = str(row[3]).strip() if row[3] else False  # Đây là Mã chi tiết (nếu có)
                employee_code = str(row[4]).strip()
                raw_salary = row[5]

                # 2. Validate cơ bản
                if not raw_month or not raw_year:
                    errors.append(f"Dòng {row_index}: Thiếu Tháng hoặc Năm")
                    continue

                try:
                    month_str = str(int(raw_month))
                    year_int = int(raw_year)
                except ValueError:
                    errors.append(f"Dòng {row_index}: Tháng/Năm lỗi định dạng")
                    continue

                if month_str not in [str(i) for i in range(1, 13)]:
                    errors.append(f"Dòng {row_index}: Tháng '{raw_month}' không hợp lệ")
                    continue

                # 3. TÌM ID CỦA LOẠI PHỤ CẤP TỪ MÃ EXCEL
                if not raw_type_code:
                    errors.append(f"Dòng {row_index}: Thiếu mã loại phụ cấp (Cột C)")
                    continue

                allowance_type_id = allowance_types_map.get(raw_type_code)

                # Nếu chưa tìm thấy trong cache, thử search lại trong DB (phòng trường hợp vừa tạo mới mà cache cũ)
                # hoặc báo lỗi luôn. Ở đây tôi chọn báo lỗi để đảm bảo dữ liệu chuẩn.
                if not allowance_type_id:
                    errors.append(
                        f"Dòng {row_index}: Mã loại phụ cấp '{raw_type_code}' chưa được khai báo trong Danh mục hệ thống.")
                    continue

                # 4. Tìm nhân viên
                employee = Employee.search([('employee_code', '=', employee_code)], limit=1)
                if not employee:
                    errors.append(f"Dòng {row_index}: Không tìm thấy NV mã '{employee_code}'")
                    continue

                # 5. Kiểm tra trùng lặp
                domain = [
                    ('employee_id', '=', employee.id),
                    ('month', '=', month_str),
                    ('year', '=', year_int),
                    ('allowance_type_id', '=', allowance_type_id),  # Thay đổi field
                    ('code', '=', raw_code)
                ]
                if not raw_code:
                    domain[-1] = ('code', '=', False)

                existing_rec = Allowance.search(domain, limit=1)

                vals = {
                    'employee_id': employee.id,
                    'month': month_str,
                    'year': year_int,
                    'allowance_type_id': allowance_type_id,  # Lưu ID relation
                    'code': raw_code,
                    'salary_allowance': float(raw_salary or 0),
                    'currency_id': self.env.company.currency_id.id
                }

                if existing_rec:
                    existing_rec.write(vals)
                else:
                    Allowance.create(vals)

                success_count += 1

            except Exception as e:
                errors.append(f"Dòng {row_index}: Lỗi hệ thống - {str(e)}")

        # --- Phần trả về kết quả giữ nguyên như cũ ---
        msg = f"Hoàn tất! Đã xử lý thành công {success_count} dòng."
        if errors:
            msg += f"\n\nCó {len(errors)} dòng lỗi:\n" + "\n".join(errors[:20])
            if len(errors) > 20:
                msg += "\n... và các lỗi khác."

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Kết quả Import',
                    'message': msg,
                    'type': 'warning',
                    'sticky': True,
                }
            }

        return {
            'effect': {
                'fadeout': 'slow',
                'message': msg,
                'type': 'rainbow_man',
            }
        }