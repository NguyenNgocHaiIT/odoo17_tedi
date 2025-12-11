# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
from openpyxl import load_workbook
from io import BytesIO


class TediInsuranceImportWizard(models.TransientModel):
    _name = 'tedi.insurance.import.wizard'
    _description = 'Import Lương BHXH từ Excel'

    file = fields.Binary(string="Tệp Excel", required=True)
    filename = fields.Char(string="Tên file")

    def action_import(self):
        if not self.file:
            raise UserError(_("Vui lòng chọn file Excel!"))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Không thể đọc file. Vui lòng kiểm tra định dạng .xlsx. Chi tiết: %s") % str(e))

        Insurance = self.env['hr.employee.insurance']
        Employee = self.env['hr.employee']

        success_count = 0
        errors = []
        row_index = 1  # Bắt đầu từ 1 (Header)

        # Cấu trúc file Excel yêu cầu:
        # Cột A (0): Month (1-12)
        # Cột B (1): Year (yyyy)
        # Cột C (2): Employee Code
        # Cột D (3): Salary BHXH

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_index += 1

            # Kiểm tra dòng trống hoặc thiếu mã nhân viên (Cột C là index 2)
            if not row or len(row) < 4 or row[2] is None:
                continue

            try:
                # 1. Lấy dữ liệu thô
                raw_month = row[0]
                raw_year = row[1]
                employee_code = str(row[2]).strip()
                raw_salary = row[3]

                # 2. Xử lý Month/Year (Excel có thể trả về float như 5.0)
                if not raw_month or not raw_year:
                    errors.append(f"Dòng {row_index}: Thiếu Tháng hoặc Năm")
                    continue

                try:
                    month_str = str(int(raw_month))  # Chuyển 5.0 -> 5 -> "5"
                    year_int = int(raw_year)
                except ValueError:
                    errors.append(f"Dòng {row_index}: Tháng/Năm phải là số")
                    continue

                if month_str not in [str(i) for i in range(1, 13)]:
                    errors.append(f"Dòng {row_index}: Tháng '{raw_month}' không hợp lệ")
                    continue

                # 3. Tìm nhân viên
                employee = Employee.search([('employee_code', '=', employee_code)], limit=1)
                if not employee:
                    errors.append(f"Dòng {row_index}: Không tìm thấy nhân viên mã '{employee_code}'")
                    continue

                # 4. Kiểm tra trùng lặp
                domain = [
                    ('employee_id', '=', employee.id),
                    ('month', '=', month_str),
                    ('year', '=', year_int)
                ]
                existing_rec = Insurance.search(domain, limit=1)

                # Logic: Nếu tồn tại -> Cập nhật (hoặc Báo lỗi tùy nghiệp vụ).
                # Ở đây tôi để code Cập nhật (Update) để tiện hơn cho người dùng,
                # nếu muốn báo lỗi như Attendance thì uncomment dòng dưới.

                vals = {
                    'employee_id': employee.id,
                    'month': month_str,
                    'year': year_int,
                    'salary_bhxh': float(raw_salary or 0),
                    'currency_id': self.env.company.currency_id.id
                }

                if existing_rec:
                    # Cách 1: Báo lỗi nếu trùng (giống Attendance)
                    # errors.append(f"Dòng {row_index}: Đã tồn tại dữ liệu tháng {month_str}/{year_int} của NV {employee_code}")
                    # continue

                    # Cách 2: Cập nhật (Khuyên dùng cho lương)
                    existing_rec.write(vals)
                else:
                    Insurance.create(vals)

                success_count += 1

            except Exception as e:
                errors.append(f"Dòng {row_index}: Lỗi xử lý - {str(e)}")

        # Kết quả trả về
        msg = f"Hoàn tất! Đã xử lý thành công {success_count} dòng."
        if errors:
            msg += f"\n\nCó {len(errors)} dòng lỗi:\n" + "\n".join(errors[:20])
            if len(errors) > 20:
                msg += "\n... và các lỗi khác."

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Kết quả Import Lương BHXH',
                    'message': msg,
                    'type': 'warning',
                    'sticky': True,
                }
            }

        return {
            'effect': {
                'fadeout': 'slow',
                'message': msg,
                'type': 'rainbow_man',
            }
        }