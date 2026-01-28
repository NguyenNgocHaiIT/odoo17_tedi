# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import openpyxl
# [MỚI] Thêm import Alignment để xử lý căn lề
from openpyxl.styles import Alignment
from datetime import datetime
import pytz
import calendar
from copy import copy


class TediImportConfig(models.Model):
    """Model lưu trữ file mẫu (Singleton)"""
    _name = 'tedi.import.config'
    _description = 'Cấu hình Import Chấm công'
    _rec_name = 'name'

    name = fields.Char(string='Tên cấu hình', default='Cấu hình mặc định', required=True)
    sample_file = fields.Binary(string='File Excel Mẫu', required=True)
    sample_filename = fields.Char(string='Tên File')

    @api.model
    def create(self, vals):
        if self.search_count([]) >= 1:
            raise UserError(_("Chỉ được phép tạo một bản ghi cấu hình. Vui lòng sửa bản ghi hiện có."))
        return super(TediImportConfig, self).create(vals)


class TediSheetSelection(models.TransientModel):
    """Model tạm để chứa danh sách sheet cho Dropdown"""
    _name = 'tedi.sheet.selection'
    _description = 'Lựa chọn Sheet Excel'
    _rec_name = 'name'

    wizard_id = fields.Many2one('tedi.attendance.import.wizard', string='Wizard')
    name = fields.Char(string='Tên Sheet', required=True)


class TediAttendanceImportWizard(models.TransientModel):
    _name = 'tedi.attendance.import.wizard'
    _description = 'Import Chấm công & Nghỉ phép'

    file = fields.Binary(string='File Excel')
    filename = fields.Char(string='Tên file')

    sheet_ids = fields.One2many('tedi.sheet.selection', 'wizard_id', string='Danh sách Sheet')
    selected_sheet_id = fields.Many2one('tedi.sheet.selection', string="Chọn Sheet dữ liệu",
                                        domain="[('id', 'in', sheet_ids)]")

    month = fields.Selection([
        ('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'), ('4', 'Tháng 4'),
        ('5', 'Tháng 5'), ('6', 'Tháng 6'), ('7', 'Tháng 7'), ('8', 'Tháng 8'),
        ('9', 'Tháng 9'), ('10', 'Tháng 10'), ('11', 'Tháng 11'), ('12', 'Tháng 12')
    ], string='Tháng', required=True, default=lambda self: str(fields.Date.today().month))

    year = fields.Integer(string='Năm', required=True, default=lambda self: fields.Date.today().year)

    first_data_row = fields.Integer(string='Dòng nhân viên đầu tiên', default=10,
                                    help="Nhập số dòng chứa dữ liệu của nhân viên đầu tiên trong file Excel")

    generated_file = fields.Binary(string='File đã tạo', readonly=True)
    generated_filename = fields.Char(string='Tên file đã tạo')

    def _get_default_sample_file(self):
        config = self.env['tedi.import.config'].search([], limit=1)
        return config.sample_file if config else False

    def _get_default_sample_filename(self):
        return "Mau_Cham_Cong_Co_Du_Lieu.xlsx"

    sample_file = fields.Binary(string='Tải File Mẫu', default=_get_default_sample_file, readonly=True)
    sample_filename = fields.Char(string='Tên File Mẫu', default=_get_default_sample_filename)

    def _copy_cell_style(self, src_cell, dst_cell):
        """Copy style từ ô nguồn sang ô đích."""
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = copy(src_cell.number_format)
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

    # ========================================================
    # ACTION: DOWNLOAD SAMPLE
    # ========================================================
    def action_download_sample(self):
        self.ensure_one()

        # --- A. Lấy cấu hình và nhân viên ---
        config = self.env['tedi.import.config'].search([], limit=1)
        if not config or not config.sample_file:
            raise UserError(_("Chưa có file mẫu trong cấu hình. Vui lòng liên hệ Admin."))

        employees = self.env['hr.contract'].search([('state', '=', 'open')]).mapped('employee_id')
        if not employees:
            raise UserError(_("Không có nhân viên nào đang có hợp đồng 'Đang chạy'."))

        num_employees = len(employees)

        # --- B. Xử lý Excel ---
        try:
            file_data = base64.b64decode(config.sample_file)
            data_file = io.BytesIO(file_data)
            wb = openpyxl.load_workbook(data_file)
            ws = wb.active

            start_row = self.first_data_row
            max_col = ws.max_column

            # --- BƯỚC 1: CHÈN DÒNG ---
            if num_employees > 1:
                ws.insert_rows(start_row + 1, amount=num_employees - 1)

            # --- GỠ BỎ MERGE TỰ ĐỘNG ---
            end_row = start_row + num_employees - 1
            merged_ranges = list(ws.merged_cells.ranges)

            for merged_cell in merged_ranges:
                if merged_cell.min_row >= start_row and merged_cell.max_row <= end_row:
                    try:
                        ws.unmerge_cells(str(merged_cell))
                    except (KeyError, ValueError, IndexError):
                        pass

            # --- CẤU HÌNH CỘT ---
            COL_STT = 1
            COL_CODE = 2
            COL_NAME = 3

            # --- BƯỚC 2: DUYỆT VÀ GHI DỮ LIỆU ---
            for i, emp in enumerate(employees):
                current_row = start_row + i

                # === COPY STYLE ===
                for col_idx in range(1, max_col + 1):
                    src_cell = ws.cell(row=start_row, column=col_idx)
                    dst_cell = ws.cell(row=current_row, column=col_idx)
                    if i > 0:
                        self._copy_cell_style(src_cell, dst_cell)

                # === GHI DỮ LIỆU ===
                # 1. Số thứ tự
                cell_stt = ws.cell(row=current_row, column=COL_STT)
                cell_stt.value = i + 1

                # [FIXED] Tạo Alignment mới thay vì sửa cái cũ để tránh lỗi Immutable
                old_align = cell_stt.alignment
                new_align = Alignment(
                    horizontal='center',  # Force center
                    vertical=old_align.vertical if old_align else None,
                    wrap_text=old_align.wrap_text if old_align else None
                )
                cell_stt.alignment = new_align

                # 2. Mã Nhân Viên
                cell_code = ws.cell(row=current_row, column=COL_CODE)
                cell_code.value = emp.employee_code or ''

                # 3. Tên Nhân Viên
                cell_name = ws.cell(row=current_row, column=COL_NAME)
                cell_name.value = emp.name

            # --- C. Lưu và Trả file ---
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            generated_data = base64.b64encode(output.read())
            wb.close()

            out_filename = f"Cham_Cong_Thang_{self.month}_{self.year}.xlsx"
            self.write({
                'generated_file': generated_data,
                'generated_filename': out_filename
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/?model=tedi.attendance.import.wizard&id=%s&field=generated_file&download=true&filename=%s' % (
                    self.id, out_filename),
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_("Lỗi xử lý file Excel: %s") % str(e))

    # --- Onchange đọc Sheet ---
    @api.onchange('file')
    def _onchange_file(self):
        self.selected_sheet_id = False
        self.sheet_ids = [(5, 0, 0)]

        if not self.file:
            return

        try:
            file_data = base64.b64decode(self.file)
            data_file = io.BytesIO(file_data)
            wb = openpyxl.load_workbook(data_file, read_only=True, keep_links=False, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            new_sheet_ids = []
            SheetModel = self.env['tedi.sheet.selection']

            for name in sheet_names:
                new_rec = SheetModel.create({'name': name})
                new_sheet_ids.append(new_rec.id)

            self.sheet_ids = [(6, 0, new_sheet_ids)]

            if new_sheet_ids:
                self.selected_sheet_id = new_sheet_ids[0]

        except Exception:
            self.sheet_ids = [(5, 0, 0)]
            self.selected_sheet_id = False

    # --- Helpers ---
    def _get_attendance_symbols(self):
        return ['+', '-']

    def _get_public_holiday_symbols(self):
        return ['L', 'TET', 'GIO_TO', 'ND', 'QP', 'LE']

    def _find_leave_type_by_code(self, symbol):
        if not symbol: return False
        clean_code = symbol.strip().upper()
        MAPPING = {'VRO': 'RO', 'CV': 'CVD', 'OM': 'O'}
        search_code = MAPPING.get(clean_code, clean_code)
        leave_type = self.env['hr.leave.type'].search([('code', '=', search_code)], limit=1)
        if not leave_type:
            leave_type = self.env['hr.leave.type'].search([('work_entry_type_id.code', '=', search_code)], limit=1)
        return leave_type.id if leave_type else False

    def _get_work_hours_from_contract(self, employee, current_date):
        contract = self.env['hr.contract'].search([
            ('employee_id', '=', employee.id),
            ('state', 'in', ['open', 'close']),
            ('date_start', '<=', current_date),
            '|', ('date_end', '=', False), ('date_end', '>=', current_date)
        ], limit=1)

        if not contract or not contract.resource_calendar_id:
            return False

        calendar = contract.resource_calendar_id

        def compute_hours_from_lines(lines):
            if not lines: return False
            h_start = min(lines.mapped('hour_from'))
            h_end = max(lines.mapped('hour_to'))

            morning_shift = lines.filtered(lambda a: a.day_period == 'morning')
            if morning_shift:
                h_noon = max(morning_shift.mapped('hour_to'))
            else:
                h_noon = h_start + (h_end - h_start) / 2
            return h_start, h_end, h_noon

        day_of_week = str(current_date.weekday())
        current_day_lines = calendar.attendance_ids.filtered(
            lambda a: a.dayofweek == day_of_week and a.display_type != 'line_section'
        )

        res = compute_hours_from_lines(current_day_lines)
        if res: return res

        monday_lines = calendar.attendance_ids.filtered(
            lambda a: a.dayofweek == '0' and a.display_type != 'line_section'
        )
        return compute_hours_from_lines(monday_lines)

    def _make_utc_from_float(self, date_obj, float_time, l_tz, u_tz):
        try:
            hours = int(float_time)
            minutes = int((float_time * 60) % 60)
            ldt = datetime(date_obj.year, date_obj.month, date_obj.day, hours, minutes)
            return l_tz.localize(ldt).astimezone(u_tz).replace(tzinfo=None)
        except:
            return False

    def _parse_to_utc(self, date_obj, t_str, l_tz, u_tz):
        try:
            t_str = t_str.replace(';', ':').replace('.', ':')[:5]
            parts = t_str.split(':')
            float_val = int(parts[0]) + int(parts[1]) / 60.0
            return self._make_utc_from_float(date_obj, float_val, l_tz, u_tz)
        except:
            return False

    # ========================================================
    # MAIN IMPORT ACTION
    # ========================================================
    def action_import(self):
        self.ensure_one()

        if not self.file: raise UserError(_("Vui lòng chọn file Excel."))
        if not self.selected_sheet_id:
            raise UserError(_("Vui lòng chọn Sheet cần import."))

        target_sheet_name = self.selected_sheet_id.name

        try:
            file_data = base64.b64decode(self.file)
            data_file = io.BytesIO(file_data)
            wb = openpyxl.load_workbook(data_file, data_only=True)
            sheet = wb[target_sheet_name]
        except Exception as e:
            raise UserError(_("Lỗi đọc file: %s") % e)

        Attendance = self.env['hr.attendance']
        Leave = self.env['hr.leave']
        Employee = self.env['hr.employee']

        ATT_SYMBOLS = self._get_attendance_symbols()
        PUB_SYMBOLS = self._get_public_holiday_symbols()
        local_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        utc_tz = pytz.utc

        COL_EMP = 1
        COL_START_DAY = 3
        START_ROW = self.first_data_row
        last_day = calendar.monthrange(self.year, int(self.month))[1]

        cnt_att = 0
        cnt_leave = 0
        cnt_skipped = 0
        errors = []
        imported_employees = set()

        for row_idx, row in enumerate(sheet.iter_rows(min_row=START_ROW, values_only=True), start=START_ROW):
            emp_code = row[COL_EMP]
            if not emp_code: continue

            emp_code_str = str(emp_code).strip()
            employee = Employee.search([('employee_code', '=', emp_code_str)], limit=1)

            if not employee:
                errors.append(f"Dòng {row_idx}: Không tìm thấy mã NV '{emp_code_str}'")
                continue

            for day in range(1, last_day + 1):
                col_idx = COL_START_DAY + (day - 1)
                if col_idx >= len(row): break

                raw_val = row[col_idx]
                if not raw_val: continue

                val = str(raw_val).strip().upper()
                try:
                    curr_date = datetime(self.year, int(self.month), day)
                except:
                    continue

                work_hours = self._get_work_hours_from_contract(employee, curr_date.date())

                if not work_hours: continue
                h_start_float, h_end_float, h_noon_float = work_hours

                is_half = True if ('/2' in val or val == '-') else False
                base_symbol = val.replace('/2', '') if '/2' in val else ('-' if val == '-' else val)

                # 1. Chấm công
                if base_symbol in ATT_SYMBOLS or ':' in val:
                    check_in_dt = False
                    if ':' in val:
                        times = val.replace('\n', ' ').split()
                        valid_times = [t for t in times if ':' in t]
                        if valid_times:
                            try:
                                check_in_dt = self._parse_to_utc(curr_date, valid_times[0], local_tz, utc_tz)
                                if len(valid_times) > 1:
                                    check_out_dt = self._parse_to_utc(curr_date, valid_times[-1], local_tz, utc_tz)
                                else:
                                    h_to_temp = h_noon_float if is_half else h_end_float
                                    check_out_dt = self._make_utc_from_float(curr_date, h_to_temp, local_tz, utc_tz)
                            except:
                                continue
                    else:
                        h_to_temp = h_noon_float if is_half else h_end_float
                        check_in_dt = self._make_utc_from_float(curr_date, h_start_float, local_tz, utc_tz)
                        check_out_dt = self._make_utc_from_float(curr_date, h_to_temp, local_tz, utc_tz)

                    if check_in_dt and check_out_dt:
                        start_d = check_in_dt.replace(hour=0, minute=0)
                        end_d = check_in_dt.replace(hour=23, minute=59)

                        exist_att = Attendance.search_count([
                            ('employee_id', '=', employee.id),
                            ('check_in', '>=', start_d), ('check_in', '<=', end_d)
                        ])

                        if not exist_att:
                            try:
                                with self.env.cr.savepoint():
                                    Attendance.create({
                                        'employee_id': employee.id,
                                        'check_in': check_in_dt,
                                        'check_out': check_out_dt,
                                        'attendance_type': 'attendance'
                                    })
                                    cnt_att += 1
                                    imported_employees.add(employee.name)
                            except Exception as e:
                                errors.append(f"Dòng {row_idx}: Lỗi tạo công - {str(e)}")

                # 2. Lễ Tết
                elif base_symbol in PUB_SYMBOLS:
                    cnt_skipped += 1
                    continue

                # 3. Nghỉ phép
                else:
                    leave_type_id = self._find_leave_type_by_code(val)
                    if not leave_type_id:
                        errors.append(f"Dòng {row_idx} ({emp_code}): Không tìm thấy loại nghỉ '{val}'")
                        continue

                    domain = [('employee_id', '=', employee.id), ('state', 'in', ['confirm', 'validate1', 'validate']),
                              ('request_date_from', '<=', curr_date.date()),
                              ('request_date_to', '>=', curr_date.date())]

                    if not Leave.search_count(domain):
                        h_to_temp = h_noon_float if is_half else h_end_float
                        dt_f = self._make_utc_from_float(curr_date, h_start_float, local_tz, utc_tz)
                        dt_t = self._make_utc_from_float(curr_date, h_to_temp, local_tz, utc_tz)

                        try:
                            with self.env.cr.savepoint():
                                vals = {
                                    'employee_id': employee.id,
                                    'holiday_status_id': int(leave_type_id),
                                    'date_from': dt_f, 'date_to': dt_t,
                                    'request_date_from': curr_date.date(),
                                    'request_date_to': curr_date.date(),
                                    'request_unit_hours': True,
                                    'request_unit_half': False,
                                    'number_of_days': 0.5 if is_half else 1,
                                    'state': 'confirm',
                                    'name': f'Import {val}',
                                    'is_imported': True
                                }

                                leave = Leave.sudo().create(vals)
                                leave_sudo = leave.with_context(bypass_manager_check=True)
                                if leave_sudo.state == 'confirm': leave_sudo.action_approve()
                                if leave_sudo.state == 'validate1': leave_sudo.action_validate()
                                if leave_sudo.state != 'validate':
                                    leave_sudo.write({'state': 'validate'})
                                    leave_sudo._validate_leave_request()

                                cnt_leave += 1
                                imported_employees.add(employee.name)

                        except Exception as ex:
                            err_str = str(ex)
                            vn_error = err_str
                            if "You do not have any allocation" in err_str:
                                vn_error = "Hết quỹ nghỉ phép."
                            elif "must be in the future" in err_str:
                                vn_error = "Ngày nghỉ phải ở tương lai."
                            elif "overlap" in err_str.lower():
                                vn_error = "Trùng thời gian với đơn khác."

                            errors.append(f"Dòng {row_idx} ({emp_code}): Lỗi tạo nghỉ '{val}' -> {vn_error}")

        msg = (f"Hoàn tất xử lý!\n"
               f"- Số nhân viên có dữ liệu: {len(imported_employees)}\n"
               f"- Tổng bản ghi Chấm công: {cnt_att}\n"
               f"- Tổng bản ghi Nghỉ Phép: {cnt_leave}\n"
               f"- Bỏ qua (Lễ/Tết): {cnt_skipped}")

        if errors:
            unique_err = list(set(errors))
            msg += f"\n\n--- CÓ LỖI ({len(unique_err)}) ---\n" + "\n".join(unique_err[:15])
            type_msg = 'warning'
        else:
            type_msg = 'success'

        return {'type': 'ir.actions.client', 'tag': 'display_notification',
                'params': {'title': 'Kết quả Import', 'message': msg, 'type': type_msg, 'sticky': True}}